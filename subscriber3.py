#!/usr/bin/env python3
import rospy
from openpyxl import Workbook
from geometry_msgs.msg import TwistStamped
from sensor_msgs.msg import Imu
from sensor_msgs.msg import NavSatFix
from firebase import firebase as fb
from queue import Queue
import scipy.integrate as spi
import folium
import math

message_queue = Queue()
data_dict = {}
# topic_imu = "kitti/oxts/imu"
# topic_gps = "/kitti/oxts/gps/fix"
topic_imu = "/imu"
topic_gps = "/gps_fix"
mapsave = '/home/trucnht/catkin_ws/src/Result/map.html'
output_xlsx = '/home/trucnht/catkin_ws/src/Result/data.xlsx'
output1_xlsx = '/home/trucnht/catkin_ws/src/Result/data1.xlsx'
output2_xlsx = '/home/trucnht/catkin_ws/src/Result/data2.xlsx'

#Coordinate interpolation function
def interpolate_points(lat1,lon1,lat2,lon2,num_points):
    d = math.sqrt( (lat2-lat1)**2 + (lon2-lon1)**2 )
    interpolation_distance = d/(num_points+1)
    lat_i = []
    lon_i = []
    for i in range (1,num_points+1):
        lat = lat1 + (lat2-lat1) * (i*interpolation_distance/d)
        lat_i.append(lat)
        lon = lon1 + (lon2-lon1) * (i*interpolation_distance/d)
        lon_i.append(lon)
    return lat_i, lon_i

#Calculation total distance traveled function
def haversine(lat1,lon1,lat2,lon2):
    R = 6371e3 #Bán kinh trái đất
    φ1 = math.radians(lat1)
    φ2 = math.radians(lat2)
    deltaφ = φ2 - φ1
    deltaλ = math.radians(lon2 - lon1)
    
    a = math.sin(deltaφ / 2) * math.sin(deltaφ / 2) + math.cos(φ1) * math.cos(φ2) * math.sin(deltaλ / 2) * math.sin(deltaλ / 2)
    c = 2 * math.atan2(math.sqrt(a),math.sqrt(1-a))
    
    return R * c

#Calcution Vertical displacement fucntion
def calculate_vertical_displacement(av, t1, t2):
    az = abs(av / 1)
    b = t2 - t1
    
    vertical_speed = az * b
    
    Vh = spi.quad(lambda t: vertical_speed, 0, b)[0]
        
    return Vh

#IRI Thresholds
thresholds = [
    (10, [(17.99, 'yellowgreen'), (32.32, 'yellow'), (float('inf'), 'red')]),
    (20, [(8.99, 'yellowgreen'), (16.16, 'yellow'), (float('inf'), 'red')]),
    (30, [(5.99, 'yellowgreen'), (10.8, 'yellow'), (float('inf'), 'red')]),
    (40, [(4.49, 'yellowgreen'), (8.08, 'yellow'), (float('inf'), 'red')]),
    (50, [(3.59, 'yellowgreen'), (6.25, 'yellow'), (float('inf'), 'red')])
]

def get_color(v, I):
    for v_min, thresholds_v in thresholds:
        if v < v_min:
            for I_min, color in thresholds_v:
                if I < I_min:
                    return color
    return 'yellowgreen'
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Get data
def callback1(msg):
    # message_queue.put(("kitti/oxts/imu", msg))
    message_queue.put((topic_imu, msg))

def callback2(msg):
    # message_queue.put(("/kitti/oxts/gps/fix", msg))
    message_queue.put((topic_gps, msg))
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Read and process data
def process_queue():
    num_points = 0
    latitude = []
    longitude = []
    velocity = []
    time_secs = []
    time_nsecs = []
    time = []
    ax = []
    ay = []
    az = []
    lat = []
    lon = []
    distances = []
    h = []
    Vh = []
    lat1 = []
    lon1 = []
    lat2 = []
    lon2 = []
    L=1.4
    t_th = []
    
    while not rospy.is_shutdown():
        if not message_queue.empty():
            topic, msg = message_queue.get()
        else:
            # Gán giá trị mặc định cho biến topic
            topic = None

        if topic == topic_gps or topic == topic_imu:
            secs = msg.header.stamp.secs
            if secs not in data_dict:
                data_dict[secs] = {
                    "latitude": None,
                    "longitude": None,
                    "imu_data": []
                }
            if topic == topic_gps:
                data_dict[secs]["latitude"] = msg.latitude
                data_dict[secs]["longitude"] = msg.longitude
                data_dict[secs]["altitude"] = msg.altitude
            elif topic == topic_imu:
                imu_data = {
                    "nsecs": msg.header.stamp.nsecs,
                    "linear_acceleration_x": msg.linear_acceleration.x,
                    "linear_acceleration_y": msg.linear_acceleration.y,
                    "linear_acceleration_z": msg.linear_acceleration.z
                }
                data_dict[secs]["imu_data"].append(imu_data)

    for secs, data in data_dict.items():
        if data["latitude"] is not None and data["longitude"] is not None and len(data["imu_data"]) > 0:
            for imu_data in data["imu_data"]:
                time_secs.append(secs)
                time_nsecs.append(imu_data["nsecs"])
                time_float = float(secs) + float(imu_data["nsecs"]) / 1e9  # Tính giá trị float từ secs và nsecs
                time.append(time_float)
                latitude.append(data["latitude"])
                longitude.append(data["longitude"])
                velocity.append(data["altitude"])
                ax.append(imu_data["linear_acceleration_x"])
                ay.append(imu_data["linear_acceleration_y"])
                az.append(imu_data["linear_acceleration_z"])

    #Coordinate interpolation 
    for i in range(len(latitude) - 1):
        if latitude[i] == latitude[i + 1] and longitude[i] == longitude[i + 1]:
            if i == 0:
                num_points += 1
                lat.append(latitude[0])
                lon.append(longitude[0])
            else:   
                num_points += 1
        else:
            lat_i, lon_i = interpolate_points(latitude[i], longitude[i], latitude[i + 1], longitude[i + 1],  num_points)
            
            for j in range(len(lat_i)):
                lat.append(round(lat_i[j], 13))
                lon.append(round(lon_i[j], 13))
            lat.append(latitude[i+1])
            lon.append(longitude[i+1])
            num_points = 0
    
    #Calculate distance traveled
    for i in range(len(lat)):
        if(i == 0):
            distances.append(0)
            Vh.append(0)
        else:
            distance = haversine(lat[i-1],lon[i-1],lat[i], lon[i])
            distances.append(distance)
            index = calculate_vertical_displacement(az[i], time[i-1], time[i])
            Vh.append(index)
            t_th.append(L/velocity[i])
    distance = haversine(lat[i-1],lon[i-1],lat[i], lon[i])
    distances.append(distance)
    index = calculate_vertical_displacement(az[i], time[i-1], time[i])
    Vh.append(index)
    t_th.append(L/velocity[i])
    total_distance = sum(distances)/1000
    print(f"S: {total_distance}")

    #Calculate IRI s = 100m
    s = 0
    s_th = 100
    vh_sum = 0
    distance_segments = []
    IRI = []
    time_start = []
    time_end = []
    for i in range(len(lat)):
        if s < s_th:
            s += distances[i]
            vh_sum += Vh[i]
            if i == 0:
                lt = lat[0]
                ln = lon[0]
                ts = time[0]
        else:
            distance_segments.append(s)
            id = vh_sum/ s_th
            IRI.append(id*(1000/s_th))
            lat1.append(lt)
            lon1.append(ln)
            lat2.append(lat[i-1])
            lon2.append(lon[i-1])
            time_start.append(ts)
            time_end.append(time[i-1])
            s = distances[i]
            lt = lat[i]
            ln = lon[i]
            ts = time[i]
            vh_sum = Vh[i]

    distance_segments.append(s)
    id = vh_sum/ s_th
    IRI.append(id*(1000/s_th))
    lat1.append(lt)
    lon1.append(ln)
    lat2.append(lat[i])
    lon2.append(lon[i])
    time_start.append(ts)
    time_end.append(time[i])

    #Detect pothole
    ts = []
    te = []  
    lat3 = []
    lon3 = [] 
    av_th = 9.8*2
    s, e, lt, ln = 0, 0, 0, 0
    i=0
    
    while i <= len(lat):
        if az[i] >= av_th:     
            lt = lat[i]
            ln = lon [i]
            s = time[i]
            e = s
            i += 1
            while i < len(lat):
                if az[i] < av_th:
                    i += 1
                else:
                    if time[i] - e < t_th[i]:
                        e = time[i]
                        i += 1
                    else:
                        if time[i] - s < t_th[i]:
                            i -= 1                  
                        else:
                            lat3.append(lt)
                            lon3.append(ln)
                            ts.append(s)
                            te.append(time[i])
                            i += 1
                        break
            if i == len(lat):
                if time[i] - s >= t_th[i]:
                    lat3.append(lt)
                    lon3.append(ln)
                    ts.append(s)
                    te.append(time[i])
                    s = 0
        else:
            i += 1
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------
    #Visualization Map               
    for i in range(len(latitude)):
        m = folium.Map(location=[latitude[0], longitude[0]], zoom_start=18, control_scale=True)    
    # for i in range(len(lat)):
    #     m = folium.Map(location=[lat[0], lon[0]], zoom_start=18, control_scale=True)    

    # Thêm lớp vệ tinh
    tile_layer_land = folium.TileLayer(
        tiles='https://mt1.google.com/vt/lyrs=m&x={x}&y={y}&z={z}',
        name='Land Map',
        attr='Map data &copy; Google',
        min_zoom=0,
        max_zoom=22).add_to(m)

    tile_layer_satellite = folium.TileLayer(
        tiles='https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}',
        name='Satellite Map',
        attr='Map data &copy; Google',
        min_zoom=0,
        max_zoom=22).add_to(m) 

    IRI1 = []
    for i in range(len(time_start)):
        start = time.index(time_start[i])
        end = time.index(time_end[i], start + 1) + 1 if time_end[i] in time[start + 1:] else len(time)
        I = IRI[i]
        if time[start] == time_start[i]:
            IRI1.extend([I] * (end - start))   
            
    for i in range(len(lat)):
        v = velocity[i]
        I = IRI1[i]
        color = get_color(v, I)
        marker = folium.CircleMarker(location=[lat[i], lon[i]], radius=1, color=color)
        # if distance_segments[i] < 95:
        #     marker = folium.CircleMarker(location=[lat[k], lon[k]], radius=1, color="yellowgreen")
        
        # Tạo popup vị trí
        popup_content = f"Latitude: {lat[i]}<br>Longitude: {lon[i]}"
        popup = folium.Popup(popup_content, max_width=200)
        
        # Thêm popup vào marker
        marker.add_child(popup)
        
        # Thêm marker vào bản đồ
        marker.add_to(m)  

    for i in range(len(lat3)): 
        line =  folium.CircleMarker(location=[lat3[i], lon3[i]], radius=1, color='black')
        popup_content = f"Latitude: {lat3[i]}<br>Longitude: {lon3[i]}<br>Time_start: {ts[i]}<br>Time_end: {te[i]}"
        popup = folium.Popup(popup_content, max_width=200)
        line.add_child(popup)  
        line.add_to(m)
    
    # Define legend HTML content
    template = """
    <div id='maplegend' class='maplegend' 
        style='position: fixed; 
            z-index:9999; 
            border:2px solid grey; 
            background-color:rgba(255, 255, 255, 1);
            border-radius:6px; 
            padding: 5px; 
            font-size:14px; 
            weight: 200px; height: min-content;
            left: 10px; 
            bottom: 50px;'>
        
    <div class='legend-title' style='color:black; weight: 600px; padding: 5px'>IRI(m/100m):</div>
    <p style='color:black;'><span style='background:yellowgreen; 
                                            opacity: 1; 
                                            width: 15px; 
                                            height: 12px; 
                                            display: inline-block;
                                            border:1px solid grey; 
                                            margin-right: 5px;'></span>Good </p>
    <p style='color:black;'><span style='background:yellow; 
                                            opacity: 1; 
                                            width: 15px; 
                                            height: 12px; 
                                            display: inline-block;
                                            border:1px solid grey; 
                                            margin-right: 5px;'></span>Fair </p>
    <p style='color:black;'><span style='background:red; 
                                            opacity: 1; 
                                            width: 15px; 
                                            height: 12px; 
                                            display: inline-block;
                                            border:1px solid grey; 
                                            margin-right: 5px;'></span>Poor </p>
    <p style='color:black;'><span style='background:black; 
                                            opacity: 1; 
                                            width: 12px; 
                                            height: 12px; 
                                            border-radius: 50%; 
                                            display: inline-block;
                                            border:1px solid grey; 
                                            margin-right: 5px;'></span>Pothole</p>
    </div>
    </div>
    """

    m.get_root().html.add_child(folium.Element(template)) 
        
    folium.LayerControl().add_to(m)
    m.add_child(tile_layer_satellite)
    # Hiển thị bản đồ
    m.save(mapsave)
    #-----------------------------------------------------------------------------------------------------------------------------------------------------------------
    #Save results to excel 
    # Tạo tệp Excel mới
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Timestamp (secs)", 
                      "Timestamp (nsecs)", 
                      "Longitude", "Latitude", 
                      "Linear acceleration x",
                      "Linear acceleration y", 
                      "Linear acceleration z", 
                      "Velocity"])

    # Ghi các giá trị đã trích xuất vào tệp Excel
    for secs, data in data_dict.items():
        if data["latitude"] is not None and data["longitude"] is not None and len(data["imu_data"]) > 0:
            for imu_data in data["imu_data"]:
                worksheet.append([secs, 
                                imu_data["nsecs"],
                                data["longitude"], data["latitude"],
                                imu_data["linear_acceleration_x"],
                                imu_data["linear_acceleration_y"], 
                                imu_data["linear_acceleration_z"],
                                data["altitude"]])
    workbook.save(output_xlsx)
    
    # Tạo tệp Excel chua các điểm tọa độ duoc noi suy 
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Timestamp (secs)", 
                      "Timestamp (nsecs)", "Timestamp",
                      "Longitude", "Latitude", 
                      "Linear acceleration x",
                      "Linear acceleration y", 
                      "Linear acceleration z", 
                      "Velocity",
                      "Distance travel","Vertical displacement",'IRI'])

    # Ghi các giá trị đã trích xuất vào tệp Excel
    for i in range (len(lat)):
                worksheet.append([time_secs[i], 
                                time_nsecs[i], time[i],
                                lon[i], lat[i],
                                ax[i], ay[i], az[i],
                                velocity[i],
                                distances[i],Vh[i],IRI1[i]])

    workbook.save(output1_xlsx)

    # Tạo tệp Excel chua các điểm tọa độ duoc noi suy 
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Time Start","Longitude", "Latitude", 
                      "Time End","Longitude", "Latitude",
                      "Distance travel","IRI"])

    # Ghi các giá trị đã trích xuất vào tệp Excel
    for i in range (len(lat1)):
                worksheet.append([time_start[i],lon1[i],lat1[i],
                                time_end[i],lon2[i], lat2[i],
                                distance_segments[i],IRI[i]])

    workbook.save(output2_xlsx)

    for i in range (len(lat)):
        data = {'timestamp': time[i], 
                'GPS': {'longitude': lon[i], 'latitude': lat[i]}, 
                'Velocity': velocity[i], 
                'IRI': IRI1[i]}
        key = f"{time_secs[i]},{str(time_nsecs[i]).zfill(9)}" 
        # firebase.post(f'/your_path/{time_secs[i]}', data)
        firebase.put('/data2/IRI', key, data)
    for i in range(len(lat3)):
        data = {
            'timestamp': te[i], 
            'GPS': {'longitude': lon3[i], 
                    'latitude': lat3[i]}
        }
        te_str = str(te[i])  # Chuyển đổi te thành chuỗi
        te_parts = te_str.split('.')  # Tách phần nguyên và phần thập phân
        key = f"{te_parts[0]},{te_parts[1]}"
        firebase.put('/data2/Pothole', key, data)
#-----------------------------------------------------------------------------------------------------------------------------------
def main():
    rospy.init_node('bag_reader_node')
    rospy.loginfo("Bag Reader Node Started")
    rospy.Subscriber(topic_imu, Imu, callback1)
    rospy.Subscriber(topic_gps, NavSatFix, callback2)
    
    process_queue()

    rospy.loginfo("Bag Reader Node Finished")


if __name__ == '__main__':
    firebase = fb.FirebaseApplication('https://mymap-e291c-default-rtdb.firebaseio.com/', authentication=None)
    main()